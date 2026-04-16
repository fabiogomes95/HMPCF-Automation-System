# =========================================================================
# SISTEMA DE BOLETIM DE ATENDIMENTO - MÓDULO DE RELATÓRIOS MENSAL
# Desenvolvido para analisar o banco SQLite e gerar relatórios em Excel.
# =========================================================================
import pandas as pd
import sqlite3
import os

# Bibliotecas do OpenPyXL usadas para ler, criar abas, pintar e bordar o Excel [1]
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def gerar_relatorio_final():
    # =====================================================================
    # 1. LOCALIZAÇÃO E LEITURA DO BANCO DE DADOS [1]
    # =====================================================================
    caminho_db = '../hospital.db'

    if not os.path.exists(caminho_db):
        print("❌ ERRO: O banco de dados 'hospital.db' não foi encontrado nesta pasta.")
        return

    try:
        conn = sqlite3.connect(caminho_db)
        
        # 1. O DISTINCT impede duplicações matemáticas
        # 2. O != '' impede que pacientes sem CPF se misturem (Efeito Cartesiano)
        query = """
        SELECT DISTINCT
            p.nome, p.dn, p.idade, p.sexo, p.raca, p.cidade, 
            a.data_atendimento, a.hora_atendimento,
            p.cpf, p.sus, p.endereco, p.numero, p.bairro, p.tel, 
            a.procedencia
        FROM pacientes p
        JOIN atendimentos a ON (p.cpf = a.cpf AND p.cpf != '') OR (p.sus = a.sus AND p.sus != '')
        """
        df = pd.read_sql_query(query, conn)
        conn.close()
        
        # O EXTERMINADOR DE CLONES: Se no banco tiver duas linhas com o exato 
        # mesmo Nome, na mesma Data e na mesma Hora, o Pandas apaga uma delas!
        df = df.drop_duplicates(subset=['nome', 'data_atendimento', 'hora_atendimento'])
        
    except Exception as e:
        print(f"❌ Erro ao ler o banco: {e}")
        return

    # =====================================================================
    # 2. INTELIGÊNCIA TEMPORAL E "REGRA DA MADRUGADA" [2]
    # =====================================================================
    # Junta a data e a hora do atendimento em uma única coluna inteligente
    df['dt_entrada'] = pd.to_datetime(
        df['data_atendimento'] + ' ' + df['hora_atendimento'], 
        errors='coerce'
    )

    # Limpa pacientes fantasmas ou sem data de entrada [2]
    df = df.dropna(subset=['dt_entrada', 'nome'])
    df = df[df['nome'].str.strip() != '']

    # A "REGRA DA MADRUGADA": Se o atendimento foi entre 00:00 e 06:59, 
    # o Python subtrai 1 dia para que ele conte no plantão do dia anterior! [2]
    df['Data_Logica'] = df['dt_entrada'].apply(
        lambda x: (x - pd.Timedelta(days=1)).date() if x.hour < 7 else x.date()
    )

    # Cria uma etiqueta de Mês/Ano (Ex: '03-2026') baseada na Data Lógica do plantão [2]
    df['Mes_Ref'] = pd.to_datetime(df['Data_Logica']).dt.strftime('%m-%Y')

    # =====================================================================
    # 3. INTERAÇÃO COM O USUÁRIO (ESCOLHA DO MÊS) [2]
    # =====================================================================
    meses = sorted(df['Mes_Ref'].unique())
    print(f"\n📅 Meses detectados no banco de dados: {meses}")
    escolha = input("👉 Digite o mês que deseja gerar (Ex: 03-2026): ").strip()

    # Pega o "pacotão" gigante do banco e filtra SÓ o mês que o usuário pediu [2]
    df_mes = df[df['Mes_Ref'] == escolha].copy()

    if df_mes.empty:
        print(f"❌ Nenhum atendimento encontrado para o mês {escolha}.")
        return

    # =====================================================================
    # 4. ORGANIZAÇÃO DOS DADOS POR PLANTÃO (DIURNO/NOTURNO) [3]
    # =====================================================================
    colunas = ['Nº', 'Nome do Paciente', 'Data Nasc.', 'Idade', 'Sexo', 'Raça/Cor', 
               'Município', 'Hora', 'CPF', 'Cartão SUS', 'Procedência', 'Endereço', 'Telefone']
    
    dados_finais = []
    
    # Ordena os atendimentos cronologicamente
    df_mes = df_mes.sort_values('dt_entrada')
    
    # Agrupa os pacientes pela Data Lógica e pelo Turno [3]
    grupos = df_mes.groupby([
        'Data_Logica', 
        df_mes['dt_entrada'].dt.hour.apply(lambda x: 'DIURNO' if 7<=x<19 else 'NOTURNO')
    ], sort=False)

    for (data, turno), grupo in grupos:
        # Cria a "Linha Cabeçalho" preta/azul do plantão [3]
        dados_finais.append({'Nº': f"PLANTÃO {turno} - {data.strftime('%d/%m/%Y')}"})
        
        # Insere os pacientes abaixo do cabeçalho
        for i, (_, r) in enumerate(grupo.iterrows(), 1):
            
            # Limpa lixos do banco (Se a procedência for 'NORMAL', a planilha fica em branco) [3]
            proc = str(r['procedencia'] if r['procedencia'] and str(r['procedencia']).lower() != 'nan' else '').strip().upper()
            if proc == 'NORMAL': proc = ''
            
            dn_br = pd.to_datetime(r['dn']).strftime('%d/%m/%Y') if r['dn'] and str(r['dn']) != 'nan' else ""
            end = f"{str(r['endereco'] or '').strip()}, {str(r['numero'] or '').strip()} - {str(r['bairro'] or '').strip()}"

            # Empacota a linha do paciente [3]
            dados_finais.append({
                'Nº': i, 'Nome do Paciente': r['nome'], 'Data Nasc.': dn_br,
                'Idade': r['idade'], 'Sexo': r['sexo'], 'Raça/Cor': r['raca'], 
                'Município': r['cidade'], 'Hora': r['dt_entrada'].strftime('%H:%M'),
                'CPF': r['cpf'], 'Cartão SUS': r['sus'], 
                'Procedência': proc, 'Endereço': end, 'Telefone': r['tel']
            })

    # =====================================================================
    # 5. MÁGICA DO EXCEL: CRIANDO ABAS SEM DELETAR O ARQUIVO
    # =====================================================================
    nome_arq = "Relatorio_Producao_BPA.xlsx"

    # Verifica se o arquivo Excel já existe na pasta
    if os.path.exists(nome_arq):
        wb = load_workbook(nome_arq) # Abre o arquivo existente
        
        # Se a aba deste mês já existir (ex: o usuário está gerando de novo para atualizar),
        # nós deletamos SÓ a aba antiga deste mês para evitar duplicatas, e recriamos vazia.
        if escolha in wb.sheetnames:
            del wb[escolha]
            
        ws = wb.create_sheet(title=escolha) # Cria a aba do mês lá no final da planilha
        print(f"🔄 Arquivo encontrado! Atualizando a aba '{escolha}'...")
    else:
        # Se o arquivo não existir (primeira vez), cria do zero
        wb = Workbook()
        ws = wb.active
        ws.title = escolha # Dá o nome do mês para a primeira aba
        print(f"✨ Arquivo novo criado com a aba '{escolha}'...")

    # =====================================================================
    # 6. ESCREVENDO E PINTANDO O EXCEL (OPENPYXL) [4]
    # =====================================================================
    # 6.1. Escreve a primeira linha (Os títulos das colunas)
    ws.append(colunas)

    # 6.2. Escreve todas as linhas de dados (Pacientes e Cabeçalhos de Plantão)
    for linha in dados_finais:
        # O .values() converte o dicionário em uma lista simples que o Excel entende
        ws.append(list(linha.values()))

    # 6.3. As Tintas e Pinceis do Excel [4]
    fill_not = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid") # Vermelho Noturno
    fill_diu = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid") # Azul Diurno
    font_h = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # 6.4. Varre a planilha linha por linha para pintar
    for row in range(1, ws.max_row + 1):
        primeira_celula = str(ws.cell(row=row, column=1).value)
        
        # Se for a linha mestre do PLANTÃO [4]
        if "PLANTÃO" in primeira_celula:
            # Mescla as células da coluna 1 (A) até a 13 (M)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
            # Decide se pinta de Vermelho ou Azul [4]
            cor = fill_not if "NOTURNO" in primeira_celula else fill_diu
            
            for col in range(1, 14):
                c = ws.cell(row=row, column=col)
                c.fill, c.font, c.alignment = cor, font_h, center
        
        # Se for um paciente comum (bota borda e centraliza algumas colunas) [4]
        else:
            for col in range(1, 14):
                c = ws.cell(row=row, column=col)
                c.border = border
                if col in [5-10]: 
                    c.alignment = center

    # 6.5. Estica as larguras das colunas para os textos não ficarem espremidos [4]
    dims = {'A':6, 'B':35, 'C':12, 'D':6, 'E':6, 'F':12, 'G':15, 'H':8, 'I':15, 'J':18, 'K':15, 'L':40, 'M':15}
    for col, width in dims.items():
        ws.column_dimensions[col].width = width

    # Congela o cabeçalho (A primeira linha nunca some ao rolar pra baixo) [4]
    ws.freeze_panes = 'A2'
    
    # Salva o arquivo permanentemente
    wb.save(nome_arq)
    print(f"✅ Relatório do mês {escolha} salvo com sucesso na planilha '{nome_arq}'!")

# Ponto de partida do Python
if __name__ == "__main__":
    gerar_relatorio_final()