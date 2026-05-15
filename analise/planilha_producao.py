"""
PLANILHA_PRODUCAO.PY — Relatório Excel de Produção Hospitalar
==============================================================
Gera planilha .xlsx com divisão DIURNO/NOTURNO pela REGRA DA MADRUGADA.

Uso direto:  python planilha_producao.py
Via Eel:     analise_gerar_relatorio_mes("04-2026")
"""

import pandas as pd
import sqlite3
import os
import sys

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from logging_setup import logger
from utils import apenas_numeros, remove_accents


def gerar_relatorio_mes(mes_ref=None):
    """
    Gera o relatório Excel para um mês específico.
    mes_ref: string no formato "MM-AAAA" (ex: "04-2026").
    Se None, lista os meses disponíveis e usa o primeiro.
    Retorna string com resultado.
    """
    caminho_db = os.path.join(os.path.dirname(__file__), '..', 'hospital.db')
    if not os.path.exists(caminho_db):
        msg = f"ERRO: Banco nao encontrado em: {caminho_db}"
        logger.error(msg)
        return msg

    try:
        conn = sqlite3.connect(caminho_db)
        query = """
            SELECT DISTINCT p.nome, p.dn, p.idade, p.sexo, p.raca,
                   p.cidade, a.data_atendimento, a.hora_atendimento,
                   p.cpf, p.sus, p.endereco, p.numero, p.bairro,
                   p.tel, a.procedencia
            FROM pacientes p
            JOIN atendimentos a ON
                (p.cpf = a.cpf AND p.cpf != '')
                OR (p.sus = a.sus AND p.sus != '')
        """
        df = pd.read_sql_query(query, conn)
        conn.close()
        df = df.drop_duplicates(subset=['nome', 'data_atendimento', 'hora_atendimento'])
    except Exception as e:
        msg = f"Erro ao ler o banco: {e}"
        logger.error(msg)
        return msg

    df['dt_entrada'] = pd.to_datetime(
        df['data_atendimento'] + ' ' + df['hora_atendimento'], errors='coerce'
    )
    df = df.dropna(subset=['dt_entrada', 'nome'])
    df = df[df['nome'].str.strip() != '']

    df['Data_Logica'] = df['dt_entrada'].apply(
        lambda x: (x - pd.Timedelta(days=1)).date() if x.hour < 7 else x.date()
    )
    df['Mes_Ref'] = pd.to_datetime(df['Data_Logica']).dt.strftime('%m-%Y')

    meses = sorted(df['Mes_Ref'].unique())
    if not meses:
        return "Nenhum mes encontrado nos dados."

    if mes_ref is None:
        mes_ref = meses[0]

    df_mes = df[df['Mes_Ref'] == mes_ref].copy()
    if df_mes.empty:
        return f"Nenhum atendimento para {mes_ref}."

    colunas = [
        'N', 'Nome do Paciente', 'Data Nasc.', 'Idade', 'Sexo',
        'Raca/Cor', 'Municipio', 'Hora', 'CPF', 'Cartao SUS',
        'Procedencia', 'Endereco', 'Telefone'
    ]
    dados_finais = []
    df_mes = df_mes.sort_values('dt_entrada')

    grupos = df_mes.groupby([
        'Data_Logica',
        df_mes['dt_entrada'].dt.hour.apply(lambda x: 'DIURNO' if 7 <= x < 19 else 'NOTURNO')
    ], sort=False)

    for (data, turno), grupo in grupos:
        dados_finais.append({'N': f"PLANTAO {turno} - {data.strftime('%d/%m/%Y')}"})
        for i, (_, r) in enumerate(grupo.iterrows(), 1):
            nome_limpo = remove_accents(r['nome'])
            cidade_limpa = remove_accents(r['cidade'])
            bairro_limpo = remove_accents(r['bairro'])
            rua_limpa = remove_accents(r['endereco'])
            cpf_f = apenas_numeros(r['cpf'])
            sus_f = apenas_numeros(r['sus'])
            proc = remove_accents(str(r['procedencia']))
            if proc == 'NORMAL':
                proc = ''
            try:
                dn_br = pd.to_datetime(r['dn']).strftime('%d/%m/%Y')
            except Exception:
                dn_br = ''
            end_f = f"{rua_limpa}, {str(r['numero'] or '').strip()} - {bairro_limpo}"
            dados_finais.append({
                'N': i, 'Nome do Paciente': nome_limpo, 'Data Nasc.': dn_br,
                'Idade': r['idade'], 'Sexo': r['sexo'], 'Raca/Cor': r['raca'],
                'Municipio': cidade_limpa, 'Hora': r['dt_entrada'].strftime('%H:%M'),
                'CPF': cpf_f, 'Cartao SUS': sus_f, 'Procedencia': proc,
                'Endereco': end_f, 'Telefone': r['tel']
            })

    nome_arq = "Relatorio_Producao_HMPCF.xlsx"
    if os.path.exists(nome_arq):
        wb = load_workbook(nome_arq)
        if mes_ref in wb.sheetnames:
            del wb[mes_ref]
        ws = wb.create_sheet(title=mes_ref)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = mes_ref

    ws.append(colunas)
    for linha in dados_finais:
        ws.append(list(linha.values()))

    fill_not = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    fill_diu = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    font_h = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for row in range(1, ws.max_row + 1):
        primeira_celula = str(ws.cell(row=row, column=1).value)
        if "PLANTAO" in primeira_celula:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
            cor = fill_not if "NOTURNO" in primeira_celula else fill_diu
            for col in range(1, 14):
                c = ws.cell(row=row, column=col)
                c.fill = cor
                c.font = font_h
                c.alignment = center
        else:
            for col in range(1, 14):
                c = ws.cell(row=row, column=col)
                c.border = border
                if col in [6, 7, 8]:
                    c.alignment = center

    dims = {'A': 6, 'B': 35, 'C': 12, 'D': 6, 'E': 6, 'F': 12,
            'G': 15, 'H': 8, 'I': 15, 'J': 18, 'K': 15, 'L': 40, 'M': 15}
    for col, width in dims.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = 'A2'
    wb.save(nome_arq)

    msg = f"Relatorio '{mes_ref}' gerado em '{nome_arq}'!"
    logger.info(msg)
    return msg


if __name__ == "__main__":
    logger.info("Meses detectados: " + str(sorted(
        pd.to_datetime(
            pd.read_sql_query(
                "SELECT DISTINCT data_atendimento FROM atendimentos",
                sqlite3.connect(os.path.join(os.path.dirname(__file__), '..', 'hospital.db'))
            )['data_atendimento'], errors='coerce'
        ).dropna().dt.strftime('%m-%Y').unique()
    )))
    mes = input("Digite o mes (Ex: 04-2026): ").strip()
    logger.info(gerar_relatorio_mes(mes))
